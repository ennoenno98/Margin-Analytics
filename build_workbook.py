"""Build a Google-Sheets-compatible margin analytics workbook.

Output: Margin_Analytics.xlsx
Upload to Google Drive -> Open with Google Sheets to activate IMPORTDATA.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

NOVADATA_URL = "https://app.novadata.io/resources/data-export-download/17715a7b-7eb5-4a11-9476-970df01c7bca"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F3864", size=12)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build():
    wb = Workbook()

    # === README ===
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Margin Analytics – Amazon DE", SECTION_FONT),
        ("", None),
        ("How to use this file:", Font(bold=True)),
        ("1. Upload this .xlsx to Google Drive.", None),
        ("2. Right-click → Open with → Google Sheets.", None),
        ("3. Go to the 'Source' tab. Cell A1 contains an IMPORTDATA() formula that "
         "pulls live data from Novadata. Allow access if Sheets asks.", None),
        ("4. The 'Dashboard' tab auto-filters to the latest period for the marketplace "
         "selected in 'Assumptions'.", None),
        ("5. Change the marketplace, target margins, or 'top sellers only' flag in 'Assumptions' "
         "to refocus the dashboard.", None),
        ("", None),
        ("Tabs:", Font(bold=True)),
        ("• Source — live Novadata pull (IMPORTDATA). Do not edit.", None),
        ("• Assumptions — marketplace, period, target CM1/CM2/CM3%, filters.", None),
        ("• Dashboard — per-SKU view for the selected marketplace + latest period.", None),
        ("• Trend — pick a SKU and see CM3% / Sales over time.", None),
        ("", None),
        ("Note: IMPORTDATA only works in Google Sheets, not Excel. If you open this in Excel, "
         "the Source tab will be empty.", Font(italic=True, color="808080")),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=text)
        if font:
            ws.cell(row=i, column=1).font = font
            if font is SECTION_FONT:
                ws.cell(row=i, column=1).fill = SECTION_FILL
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # === Source ===
    ws = wb.create_sheet("Source")
    ws["A1"] = f'=IMPORTDATA("{NOVADATA_URL}")'
    # Provide a column reference card in row 50+ so users know the schema (won't conflict
    # with IMPORTDATA which expands down from A1)
    ws.column_dimensions["A"].width = 12
    for col_letter, width in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", [12]*26):
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["I"].width = 40  # Product
    ws.column_dimensions["B"].width = 18  # Seller Partner ID

    # === Assumptions ===
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "Assumptions"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:C1")

    rows = [
        ("Marketplace", "amazon.de", "amazon.de | amazon.co.uk | amazon.fr | amazon.it | amazon.es | amazon.nl | amazon.ie | amazon.com.be | amazon.se"),
        ("Period (latest available)", '=IFERROR(MAX(Source!A2:A),"")', "Auto: max date found in Source."),
        ("Target CM1 %", 71.0, "Gross margin target after COGS only (percent points, e.g. 71 = 71%)."),
        ("Target CM2 %", 33.2, "Margin target after Amazon fees."),
        ("Target CM3 %", 19.7, "Channel margin target after marketing."),
        ("Top sellers only?", "no", "yes / no — limit Dashboard to SKUs flagged 'Top Seller' for the selected marketplace."),
        ("Min Days of Supply alert", 30, "Highlight SKUs in Dashboard with Days of Supply below this."),
    ]
    for i, (label, value, hint) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=value)
        c.fill = INPUT_FILL
        c.border = BORDER
        ws.cell(row=i, column=3, value=hint).font = Font(italic=True, color="808080")

    # Named-cell convenience: marketplace=B3, period=B4, target_cm1=B5, target_cm2=B6, target_cm3=B7,
    # top_only=B8, min_dos=B9

    # === Dashboard ===
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 16  # SKU
    ws.column_dimensions["B"].width = 50  # Product
    for col_letter, width in zip("CDEFGHIJKLMNOPQ", [10,8,8,11,9,9,9,10,12,7,7,10,10,11]):
        ws.column_dimensions[col_letter].width = width

    ws["A1"] = "Margin Dashboard"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:Q1")

    ws["A2"] = "Marketplace:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "=Assumptions!B3"
    ws["D2"] = "Period:"
    ws["D2"].font = Font(bold=True)
    ws["E2"] = "=Assumptions!B4"
    ws["G2"] = "Target CM3%:"
    ws["G2"].font = Font(bold=True)
    ws["H2"] = "=Assumptions!B7"

    headers = [
        "SKU", "Product", "Top Seller", "Orders", "Units", "Sales (€)",
        "CM1 %", "CM2 %", "CM3 %", "Δ CM3 vs target",
        "Ad Spend (€)", "ROAS", "CTR %",
        "FBA Avail", "Days Supply", "Sales Velocity", "ASIN",
    ]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header(ws, HEADER_ROW, len(headers))
    ws.row_dimensions[HEADER_ROW].height = 32
    ws.freeze_panes = "A5"

    # Filter source rows by marketplace + latest period (+ top seller flag).
    # The "Top Seller" filter column depends on marketplace; emulate with nested IF.
    # Source columns:
    #  A Period, B Seller Partner ID, C Store Name, D Marketplace, E Marketplace Name,
    #  F Parent ASIN, G Child ASIN, H SKU, I Product, J CM1%, K CM2%, L CM3%,
    #  M Sponsored Spend, N ROAS, O CTR, P Orders, Q Units, R Product Sales,
    #  S FBA Available, T Days of Supply, U Sales Velocity, V Brand,
    #  W UK filters, X FR filters, Y ES filters, Z DE filters, AA IT filters

    # Pick the right filter column based on marketplace name (only DE/UK/FR/ES/IT have flags;
    # others get blank, which won't match "Top Seller" so the top-sellers filter just empties).
    top_col = (
        'CHOOSE(MATCH(Assumptions!B3,'
        '{"amazon.de","amazon.co.uk","amazon.fr","amazon.es","amazon.it"},0),'
        'Source!Z:Z,Source!W:W,Source!X:X,Source!Y:Y,Source!AA:AA)'
    )

    # FILTER returns the SKU column; we then horizontally pull related fields with multiple FILTER
    # calls sharing the same condition. Sheets evaluates these as arrays.
    base_cond = (
        '(Source!E2:E=Assumptions!B3)*'
        '(Source!A2:A=Assumptions!B4)*'
        f'(IF(LOWER(Assumptions!B8)="yes",{top_col.replace(":Z",":Z").replace(":W",":W").replace(":X",":X").replace(":Y",":Y").replace(":AA",":AA").replace("Source!Z","Source!Z2:Z").replace("Source!W","Source!W2:W").replace("Source!X","Source!X2:X").replace("Source!Y","Source!Y2:Y").replace("Source!AA","Source!AA2:AA")}="Top Seller",1))'
    )
    # The above is messy; rebuild cleanly:
    top_range = (
        'CHOOSE(MATCH(Assumptions!B3,'
        '{"amazon.de","amazon.co.uk","amazon.fr","amazon.es","amazon.it"},0),'
        'Source!Z2:Z,Source!W2:W,Source!X2:X,Source!Y2:Y,Source!AA2:AA)'
    )
    cond = (
        f'(Source!E2:E=Assumptions!B3)*'
        f'(Source!A2:A=Assumptions!B4)*'
        f'(IF(LOWER(Assumptions!B8)="yes",{top_range}="Top Seller",1))'
    )

    def filter_col(src_range):
        # Returns a FILTER() that emits the given Source column for matching rows.
        return f'IFERROR(FILTER({src_range},{cond}),"")'

    # Row 5 onward — write FILTER formulas in row 5. In Sheets each FILTER expands its array
    # downward, so we only need to write row 5.
    formulas = {
        1: filter_col("Source!H2:H"),          # SKU
        2: filter_col("Source!I2:I"),          # Product
        3: filter_col(top_range),              # Top Seller flag for selected marketplace
        4: filter_col("Source!P2:P"),          # Orders
        5: filter_col("Source!Q2:Q"),          # Units
        6: filter_col("Source!R2:R"),          # Product Sales
        7: filter_col("Source!J2:J"),          # CM1%
        8: filter_col("Source!K2:K"),          # CM2%
        9: filter_col("Source!L2:L"),          # CM3%
        # Δ CM3 vs target — leave as a separate ARRAYFORMULA in col 10
        11: filter_col("Source!M2:M"),         # Sponsored Spend
        12: filter_col("Source!N2:N"),         # ROAS
        13: filter_col("Source!O2:O"),         # CTR
        14: filter_col("Source!S2:S"),         # FBA Available
        15: filter_col("Source!T2:T"),         # Days of Supply
        16: filter_col("Source!U2:U"),         # Sales Velocity
        17: filter_col("Source!G2:G"),         # Child ASIN
    }
    for col, formula in formulas.items():
        ws.cell(row=5, column=col, value="=" + formula)

    # Δ CM3 column — derived from col I (CM3 %). Use ARRAYFORMULA bound to a long range
    # so values appear for every row the FILTER emits.
    ws.cell(row=5, column=10,
            value='=ARRAYFORMULA(IF(LEN(I5:I)=0,,IFERROR(I5:I*1-Assumptions!B7*1,"")))')

    # Conditional formatting: highlight CM3% below target red, above target green.
    cm3_range = f"I5:I500"
    delta_range = f"J5:J500"
    ws.conditional_formatting.add(
        cm3_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(I5),I5<Assumptions!$B$7)'],
            fill=BAD_FILL,
            stopIfTrue=False,
        ),
    )
    ws.conditional_formatting.add(
        cm3_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(I5),I5>=Assumptions!$B$7)'],
            fill=GOOD_FILL,
            stopIfTrue=False,
        ),
    )
    # Days of Supply alert
    dos_range = "O5:O500"
    ws.conditional_formatting.add(
        dos_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(O5),O5<Assumptions!$B$9)'],
            fill=BAD_FILL,
            stopIfTrue=False,
        ),
    )

    # === Trend ===
    ws = wb.create_sheet("Trend")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "SKU Trend"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    ws.merge_cells("A1:E1")

    ws["A3"] = "SKU to inspect"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "VV-VITA-50"
    ws["B3"].fill = INPUT_FILL
    ws["B3"].border = BORDER
    ws["C3"] = "Type a SKU exactly as it appears in Source (e.g. VV-VITA-50)."
    ws["C3"].font = Font(italic=True, color="808080")

    ws["A4"] = "Marketplace"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = "=Assumptions!B3"

    headers = ["Period", "CM1 %", "CM2 %", "CM3 %", "Sales (€)", "Orders", "Ad Spend (€)", "ROAS"]
    HEADER_ROW = 6
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header(ws, HEADER_ROW, len(headers))
    ws.freeze_panes = "A7"

    cond_t = '(Source!H2:H=$B$3)*(Source!E2:E=$B$4)'
    pairs = [
        (1, "Source!A2:A"),  # Period
        (2, "Source!J2:J"),  # CM1
        (3, "Source!K2:K"),  # CM2
        (4, "Source!L2:L"),  # CM3
        (5, "Source!R2:R"),  # Sales
        (6, "Source!P2:P"),  # Orders
        (7, "Source!M2:M"),  # Spend
        (8, "Source!N2:N"),  # ROAS
    ]
    for col, src in pairs:
        ws.cell(row=7, column=col, value=f'=IFERROR(SORT(FILTER({src},{cond_t}),'
                                          f'FILTER(Source!A2:A,{cond_t}),TRUE),"")'
                                          if col != 1 else
                                          f'=IFERROR(SORT(FILTER({src},{cond_t})),"")')

    # === Save ===
    out_path = "Margin_Analytics.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    build()
