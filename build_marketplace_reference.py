"""Build a clean, lookup-friendly marketplace P&L reference workbook.

Reads the per-marketplace P&L exports in ``pl_exports/`` (one .xlsx per
Amazon marketplace, with monthly columns + a Total column) and writes a
single reference workbook that is easy to pull from another Excel file
with XLOOKUP / INDEX-MATCH / SUMIFS.

Output: Marketplace_PL_by_Country_Month.xlsx

Sheets
------
README              What this is, the country mapping, sign/unit conventions,
                    and copy-paste lookup formulas.
Summary             Headline KPIs per country for the whole period (readable
                    P&L laid out with metrics as rows, countries as columns),
                    plus an All-Marketplaces column.
By Country & Month  Flat table — one row per Country x Month, every P&L line
                    item as a column. Built for XLOOKUP on the "Key" column.
Data (Tidy)         Long format — Country / Month / Metric / Value, one row per
                    data point. Built for SUMIFS and PivotTables.

The per-marketplace files carry no country label, so each file was matched
to its marketplace by comparing its Units / Sales totals against the
Novadata margin export (near-exact match on both, see README).
"""
from __future__ import annotations

import datetime as dt
import glob
import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- Source files, in descending-sales (export) order ----------------------
# (filename in pl_exports/, marketplace, country, ISO code)
MARKETPLACES = [
    ("pl_2026-06-19_amazon.de.xlsx",     "amazon.de",     "Germany",     "DE"),
    ("pl_2026-06-19_amazon.it.xlsx",     "amazon.it",     "Italy",       "IT"),
    ("pl_2026-06-19_amazon.es.xlsx",     "amazon.es",     "Spain",       "ES"),
    ("pl_2026-06-19_amazon.fr.xlsx",     "amazon.fr",     "France",      "FR"),
    ("pl_2026-06-19_amazon.ie.xlsx",     "amazon.ie",     "Ireland",     "IE"),
    ("pl_2026-06-19_amazon.nl.xlsx",     "amazon.nl",     "Netherlands", "NL"),
    ("pl_2026-06-19_amazon.com.be.xlsx", "amazon.com.be", "Belgium",     "BE"),
]
SRC_DIR = "pl_exports"
OUT = "Marketplace_PL_by_Country_Month.xlsx"
AS_OF = "2026-06-19"          # latest data date (June is therefore partial)
ALL_LABEL = "All Marketplaces"

# --- Styling ----------------------------------------------------------------
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color=NAVY, size=12)
TITLE_FONT = Font(bold=True, color=NAVY, size=16)
RESULT_FILL = PatternFill("solid", fgColor="EAF1FB")
TOTALCOL_FILL = PatternFill("solid", fgColor="FFF2CC")
GREY = Font(italic=True, color="808080")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EUR_FMT = "#,##0;[Red]-#,##0"
EUR_FMT2 = "#,##0.00;[Red]-#,##0.00"
CNT_FMT = "#,##0"
PCT_FMT = '0.0"%"'

# Headline lines for the Summary sheet (clean metric names, in P&L order)
HEADLINE = [
    "Units", "PPC Units", "Returned Units", "Returned Units %",
    "__rule__",
    "Sales", "Other Sales", "Sales deductions", "Net Sales", "Net Sales %",
    "__rule__",
    "Cost of goods", "Gross Profit", "Gross Margin",
    "__rule__",
    "Amazon fees", "Profit before Ads", "Margin before Ads",
    "__rule__",
    "PPC Spend", "Net Profit", "Net Margin", "TACOS %", "ACOS %",
]
# Bold "result" lines (subtotals) in the Summary
RESULT_LINES = {"Net Sales", "Gross Profit", "Profit before Ads", "Net Profit"}


# --- Parsing helpers --------------------------------------------------------
def clean_metric(raw: str):
    """Return (clean_name, unit, level) for a raw P&L category label."""
    level = 1 if "└" in raw or "└" in raw else 0
    name = raw.replace("└", "").replace("└", "").strip()
    unit = "count"
    if name.endswith("(%)"):
        unit = "%"
        name = name[:-3].strip()
    elif name.endswith("(€)"):
        unit = "€"
        name = name[:-3].strip()
    return name, unit, level


def parse_num(v):
    """Parse a P&L cell into a float (or None if blank)."""
    if v is None:
        return None
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", "")
    if s in ("", "-", "–", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_period(header_cell):
    """'Apr 26' -> ('Apr 2026', date(2026,4,1)); 'Total' -> ('Total', None)."""
    s = str(header_cell).strip()
    if s.lower() == "total":
        return "Total", None
    d = dt.datetime.strptime(s, "%b %y").date()
    return d.strftime("%b %Y"), d


GROUP_ORDER = ["Volume", "Sales", "Cost of Goods", "Amazon Fees",
               "Advertising", "Subscriptions", "Ratios", "Other"]


def build_schema(paths):
    """Fix metric order/level/unit/group from the union of all P&L files.

    Most files are a subset of the largest marketplace, but a few small fee
    lines only appear in some countries, so we merge every file. Ordering is
    by P&L group, then first-seen, which keeps a logical column layout even
    for the extra lines.
    """
    block_groups = ["Volume", "Sales", "Cost of Goods", "Amazon Fees",
                    "Advertising", "Subscriptions", "Subscriptions"]
    ratio_names = {"Returned Units %", "TACOS %", "ACOS %"}

    schema, seen_idx = {}, {}
    counter = 0
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["P&L"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        block = 0
        for r in rows[1:]:                  # skip the 'Category' header row
            raw = r[0]
            if raw is None or str(raw).strip() == "":
                block += 1
                continue
            name, unit, level = clean_metric(str(raw))
            group = block_groups[block] if block < len(block_groups) else "Other"
            if name in ratio_names:
                group = "Ratios"
            if name not in schema:
                schema[name] = {"unit": unit, "level": level, "group": group}
                seen_idx[name] = counter
                counter += 1

    order = sorted(schema, key=lambda n: (GROUP_ORDER.index(schema[n]["group"]),
                                          seen_idx[n]))
    return schema, order


def read_pl(path):
    """Return {period_label: {metric_clean: value}} and the period metadata."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["P&L"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    periods = [parse_period(c) for c in rows[0][1:]]        # skip 'Category'
    data = {label: {} for label, _ in periods}
    for r in rows[1:]:
        raw = r[0]
        if raw is None or str(raw).strip() == "":
            continue
        name, _, _ = clean_metric(str(raw))
        for i, (label, _) in enumerate(periods, start=1):
            val = parse_num(r[i]) if i < len(r) else None
            data[label][name] = val
    return data, periods


def aggregate(per_country, period_labels, order):
    """Build the All-Marketplaces figures for each period.

    Additive €/count lines are summed; ratios are recomputed from the
    summed bases (ACOS is not derivable from this export, so left blank).
    """
    out = {label: {} for label in period_labels}

    def s(label, metric):
        tot = 0.0
        for c in per_country.values():
            v = c["data"].get(label, {}).get(metric)
            if v is not None:
                tot += v
        return tot

    for label in period_labels:
        for name in order:
            unit = SCHEMA[name]["unit"]
            if unit in ("€", "count"):
                out[label][name] = s(label, name)

        net_sales = s(label, "Net Sales")
        gross_base = s(label, "Sales") + s(label, "Other Sales")
        units = s(label, "Units")

        def pct(num, den):
            return (num / den * 100.0) if den else None

        out[label]["Net Sales %"] = pct(net_sales, gross_base)
        out[label]["Gross Margin"] = pct(s(label, "Gross Profit"), net_sales)
        out[label]["Margin before Ads"] = pct(s(label, "Profit before Ads"), net_sales)
        out[label]["Net Margin"] = pct(s(label, "Net Profit"), net_sales)
        out[label]["Returned Units %"] = pct(s(label, "Returned Units"), units)
        out[label]["TACOS %"] = pct(abs(s(label, "PPC Spend")), s(label, "Sales"))
        out[label]["ACOS %"] = None        # needs ad-attributed sales (not in export)
        if "Sales penetration from Save & Subscribe" in SCHEMA:
            out[label]["Sales penetration from Save & Subscribe"] = pct(
                s(label, "S&S Sales"), net_sales)
    return out


def num_format(unit):
    return {"€": EUR_FMT2, "%": PCT_FMT}.get(unit, CNT_FMT)


# --- Sheet builders ---------------------------------------------------------
def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER


def build_readme(wb, period_labels, month_labels):
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 95

    def put(r, text, font=None, fill=None, col="B", span=None):
        cell = ws.cell(row=r, column=2 if col == "B" else 3, value=text)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if span:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        return cell

    r = 2
    put(r, "Marketplace P&L — by Country & Month", TITLE_FONT, span=3); r += 2
    put(r, f"Period: {month_labels[0]} – {month_labels[-1]}   •   "
           f"As of {AS_OF}  (June is a partial month)   •   Currency: EUR (€)",
        GREY, span=3); r += 2

    put(r, "What this is", SECTION_FONT, SECTION_FILL, span=3); r += 1
    for line in [
        "A reference workbook built from the seven per-marketplace Amazon P&L "
        "exports (one .xlsx per marketplace) in the pl_exports/ folder.",
        "Three ready-to-reference layouts of the same numbers — pick whichever "
        "suits the formula you want to write in your other workbook.",
    ]:
        put(r, "•"); put(r, line, col="C"); r += 1
    r += 1

    put(r, "Sheets", SECTION_FONT, SECTION_FILL, span=3); r += 1
    for name, desc in [
        ("Summary", "Headline P&L per country for the whole period (metrics as "
                    "rows). Quick read; includes an All-Marketplaces column."),
        ("By Country & Month", "Flat table: one row per Country × Month, every "
                    "line item as a column. Best for XLOOKUP / INDEX-MATCH."),
        ("Data (Tidy)", "Long format: one row per Country × Month × Metric. "
                    "Best for SUMIFS and PivotTables."),
    ]:
        put(r, name, Font(bold=True)); put(r, desc, col="C"); r += 1
    r += 1

    put(r, "Country mapping", SECTION_FONT, SECTION_FILL, span=3); r += 1
    put(r, "The exports carry no country label. Each file was matched to its "
           "marketplace by its Units & Sales totals against the Novadata margin "
           "export — a near-exact match on both (within ~1%).", col="C"); r += 1
    hdr_r = r
    for j, h in enumerate(["Country", "Marketplace", "Source file"]):
        ws.cell(row=r, column=2 + j, value=h)
    ws.column_dimensions["D"].width = 34
    style_header_row(ws, r, 3, start_col=2); r += 1
    for fn, mk, country, code in MARKETPLACES:
        ws.cell(row=r, column=2, value=f"{country} ({code})").border = BORDER
        ws.cell(row=r, column=3, value=mk).border = BORDER
        ws.cell(row=r, column=4, value=f"{SRC_DIR}/{fn}").border = BORDER
        r += 1
    ws.cell(row=r, column=2, value=ALL_LABEL).border = BORDER
    ws.cell(row=r, column=3, value="sum of the 7 above").border = BORDER
    ws.cell(row=r, column=4, value="(computed)").border = BORDER
    r += 2

    put(r, "Conventions", SECTION_FONT, SECTION_FILL, span=3); r += 1
    for line in [
        "Signs follow the Amazon P&L: revenue is positive; costs, fees, ad "
        "spend, refunds and VAT are negative. A column therefore sums down to "
        "Net Profit.",
        "Percent rows are stored as points (71.0 means 71%), shown with a "
        "trailing %. Reference the cell directly — no need to divide by 100.",
        "All values are EUR; UK (GBP) and Sweden (SEK) marketplaces are not in "
        "this export.",
        "All-Marketplaces margins are recomputed from the summed €/unit bases. "
        "ACOS has no all-up value (it needs ad-attributed sales, not in this "
        "export) and is left blank.",
    ]:
        put(r, "•"); put(r, line, col="C"); r += 1
    r += 1

    put(r, "Using it from another Excel", SECTION_FONT, SECTION_FILL, span=3); r += 1
    for line in [
        'XLOOKUP a single cell — Germany, May, Net Profit:',
        '   =XLOOKUP("Germany | May 2026",'
        "'By Country & Month'!$A:$A,'By Country & Month'!$X:$X)",
        '   (point the result column at the metric you need.)',
        'SUMIFS from the tidy sheet — Spain, Total, Gross Profit:',
        "   =SUMIFS('Data (Tidy)'!$J:$J,'Data (Tidy)'!$A:$A,\"Spain\","
        "'Data (Tidy)'!$E:$E,\"Gross Profit\",'Data (Tidy)'!$H:$H,\"Total\")",
        'The "Key" column (Country | Month) on the flat sheet is there to make '
        "XLOOKUP a one-liner.",
    ]:
        put(r, "•" if not line.startswith(" ") else "")
        put(r, line, GREY if line.startswith(" ") else None, col="C"); r += 1


def build_summary(wb, per_country, agg, period_labels, month_labels):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    countries = [c["country"] for c in per_country.values()] + [ALL_LABEL]
    ncols = 1 + len(countries)

    ws.cell(row=1, column=1, value="Marketplace P&L — Summary").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sub = ws.cell(row=2, column=1,
                  value=f"Whole period total ({month_labels[0]} – "
                        f"{month_labels[-1]}, as of {AS_OF}; June partial). "
                        f"Values in EUR; % rows in points.")
    sub.font = GREY
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    hr = 4
    ws.cell(row=hr, column=1, value="P&L line (period total)")
    for j, country in enumerate(countries, start=2):
        ws.cell(row=hr, column=j, value=country)
    style_header_row(ws, hr, ncols)
    ws.row_dimensions[hr].height = 30

    ws.column_dimensions["A"].width = 26
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15

    src = list(per_country.values())
    r = hr + 1
    for metric in HEADLINE:
        if metric == "__rule__":
            r += 1
            continue
        unit = SCHEMA.get(metric, {}).get("unit", "count")
        # whole euros read cleaner in the headline summary
        fmt = {"€": EUR_FMT, "%": PCT_FMT}.get(unit, CNT_FMT)
        disp = {"Gross Margin": "Gross Margin %", "Net Margin": "Net Margin %",
                "Margin before Ads": "Margin before Ads %"}.get(metric, metric)
        is_result = metric in RESULT_LINES
        lc = ws.cell(row=r, column=1, value=disp)
        lc.font = Font(bold=True) if is_result else Font()
        for j, c in enumerate(src, start=2):
            v = c["data"].get("Total", {}).get(metric)
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = fmt
            if is_result:
                cell.font = Font(bold=True)
        v = agg.get("Total", {}).get(metric)
        cell = ws.cell(row=r, column=ncols, value=v)
        cell.number_format = fmt
        cell.font = Font(bold=True)
        if is_result:
            for j in range(1, ncols + 1):
                ws.cell(row=r, column=j).fill = RESULT_FILL
        else:
            ws.cell(row=r, column=ncols).fill = TOTALCOL_FILL
        r += 1


def build_wide(wb, per_country, agg, period_labels, order):
    ws = wb.create_sheet("By Country & Month")
    ws.sheet_view.showGridLines = False

    fixed = ["Key", "Country", "Marketplace", "Code", "Month", "Period start"]

    def col_header(n):
        unit = SCHEMA[n]["unit"]
        if unit not in ("€", "%"):
            return n
        if unit == "%" and n.rstrip().endswith("%"):
            return n                        # avoid an awkward "TACOS % (%)"
        return f"{n} ({unit})"

    metric_cols = [(col_header(n), n) for n in order]
    headers = fixed + [h for h, _ in metric_cols]

    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 42

    period_start = {}
    _, periods = read_pl(os.path.join(SRC_DIR, MARKETPLACES[0][0]))
    for label, d in periods:
        period_start[label] = d

    rows_src = [(c["country"], c["marketplace"], c["code"], c["data"])
                for c in per_country.values()]
    rows_src.append((ALL_LABEL, "(sum)", "ALL", agg))

    r = 2
    for country, mk, code, data in rows_src:
        for label, _ in periods:
            key = f"{country} | {label}"
            ws.cell(row=r, column=1, value=key)
            ws.cell(row=r, column=2, value=country)
            ws.cell(row=r, column=3, value=mk)
            ws.cell(row=r, column=4, value=code)
            ws.cell(row=r, column=5, value=label)
            ps = period_start.get(label)
            psc = ws.cell(row=r, column=6, value=ps)
            if ps:
                psc.number_format = "yyyy-mm-dd"
            for k, (_, name) in enumerate(metric_cols, start=7):
                v = data.get(label, {}).get(name)
                cell = ws.cell(row=r, column=k, value=v)
                cell.number_format = num_format(SCHEMA[name]["unit"])
            if label == "Total":
                for j in range(1, len(headers) + 1):
                    ws.cell(row=r, column=j).fill = TOTALCOL_FILL
            r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 13
    for j in range(7, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = "G2"

    ref = f"A1:{get_column_letter(len(headers))}{r - 1}"
    tbl = Table(displayName="PL_ByCountryMonth", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


def build_tidy(wb, per_country, agg, period_labels, order):
    ws = wb.create_sheet("Data (Tidy)")
    ws.sheet_view.showGridLines = False

    headers = ["Country", "Marketplace", "Code", "Group", "Metric", "Unit",
               "Level", "Month", "Period start", "Value"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))

    _, periods = read_pl(os.path.join(SRC_DIR, MARKETPLACES[0][0]))

    rows_src = [(c["country"], c["marketplace"], c["code"], c["data"])
                for c in per_country.values()]
    rows_src.append((ALL_LABEL, "(sum)", "ALL", agg))

    r = 2
    for country, mk, code, data in rows_src:
        for name in order:
            meta = SCHEMA[name]
            for label, d in periods:
                v = data.get(label, {}).get(name)
                if v is None and country == ALL_LABEL:
                    continue
                ws.cell(row=r, column=1, value=country)
                ws.cell(row=r, column=2, value=mk)
                ws.cell(row=r, column=3, value=code)
                ws.cell(row=r, column=4, value=meta["group"])
                ws.cell(row=r, column=5, value=name)
                ws.cell(row=r, column=6, value=meta["unit"])
                ws.cell(row=r, column=7, value=meta["level"])
                ws.cell(row=r, column=8, value=label)
                psc = ws.cell(row=r, column=9, value=d)
                if d:
                    psc.number_format = "yyyy-mm-dd"
                vc = ws.cell(row=r, column=10, value=v)
                vc.number_format = num_format(meta["unit"])
                r += 1

    widths = [13, 14, 6, 15, 34, 6, 6, 11, 13, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    ref = f"A1:{get_column_letter(len(headers))}{r - 1}"
    tbl = Table(displayName="PL_Tidy", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)


# --- Main -------------------------------------------------------------------
SCHEMA, ORDER = build_schema([os.path.join(SRC_DIR, fn)
                              for fn, *_ in MARKETPLACES])


def main():
    per_country = {}
    period_labels = None
    month_labels = None
    for fn, mk, country, code in MARKETPLACES:
        data, periods = read_pl(os.path.join(SRC_DIR, fn))
        per_country[country] = {"marketplace": mk, "country": country,
                                "code": code, "data": data}
        if period_labels is None:
            period_labels = [lbl for lbl, _ in periods]
            month_labels = [lbl for lbl, d in periods if d is not None]

    agg = aggregate(per_country, period_labels, ORDER)

    wb = Workbook()
    build_readme(wb, period_labels, month_labels)
    build_summary(wb, per_country, agg, period_labels, month_labels)
    build_wide(wb, per_country, agg, period_labels, ORDER)
    build_tidy(wb, per_country, agg, period_labels, ORDER)
    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(per_country)} marketplaces, "
          f"{len(ORDER)} line items, periods={period_labels})")


if __name__ == "__main__":
    main()
